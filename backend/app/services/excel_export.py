"""
Excel export service — accounting export format (legacy spreadsheet layout).
"""
import asyncio
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from typing import Any


THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
BLUE_FILL = PatternFill("solid", fgColor="CFE2F3")


def _apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = BORDER


def _bold_center(cell, value, fill=None):
    cell.value = value
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    cell.border = BORDER


async def build_accounting_excel(rows: list, date_from: date, date_to: date, filepath: str):
    """
    Build ЗП-ОТД Excel: worker name | total hours per period.
    Matches the legacy spreadsheet column layout used for accounting exports.
    """
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _build_accounting_sync, rows, date_from, date_to, filepath)


def _build_accounting_sync(rows, date_from, date_to, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "ЗП-ОТД"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15

    ws["A1"] = "Начальная дата начисления ЗП"
    ws["B1"] = str(date_from)
    ws["A2"] = "Конечная дата начисления ЗП"
    ws["B2"] = str(date_to)

    for cell in [ws["A1"], ws["B1"], ws["A2"], ws["B2"]]:
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    _bold_center(ws["A3"], "Имя сотрудника", HEADER_FILL)
    _bold_center(ws["B3"], "Итого часов", HEADER_FILL)

    hours_by_user: dict[str, float] = defaultdict(float)
    for report, user in rows:
        name = (user.full_name if user else None) or report.reg_name or f"ID:{report.user_id}"
        hours_by_user[name] += float(report.hours or 0)

    row_num = 4
    total = 0.0
    for name, hrs in sorted(hours_by_user.items()):
        ws.cell(row=row_num, column=1, value=name).border = BORDER
        hrs_cell = ws.cell(row=row_num, column=2, value=round(hrs, 2))
        hrs_cell.border = BORDER
        hrs_cell.alignment = Alignment(horizontal="center")
        total += hrs
        row_num += 1

    total_row = row_num
    ws.cell(total_row, 1, "итого").font = Font(bold=True)
    ws.cell(total_row, 1).border = BORDER
    ws.cell(total_row, 1).fill = BLUE_FILL
    ws.cell(total_row, 2, round(total, 2)).font = Font(bold=True)
    ws.cell(total_row, 2).border = BORDER
    ws.cell(total_row, 2).fill = BLUE_FILL
    ws.cell(total_row, 2).alignment = Alignment(horizontal="center")

    wb.save(filepath)


async def build_otd_excel(reports: list, filepath: str):
    """Build full OTD report Excel with all 11 columns."""
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _build_otd_sync, reports, filepath)


def _build_otd_sync(reports, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "ОТД Отчёты"

    headers = [
        "ID пользователя", "Имя", "Дата работы", "Часы",
        "Тип работ", "Тип Техники", "Техника",
        "Вид деятельности", "Вид работы", "Поле", "Культура"
    ]
    col_widths = [15, 25, 15, 8, 12, 14, 20, 20, 20, 20, 15]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, i, h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 40

    for row_idx, report in enumerate(reports, 2):
        grp = report.activity_grp or ""
        tech_type = report.machine_type if grp == "техника" else ""
        tech_name = report.machine_name if grp == "техника" else ""
        activity = report.activity if grp == "техника" else ""
        hand_work = report.activity if grp == "ручная" else ""

        values = [
            report.user_id,
            report.reg_name,
            str(report.work_date) if report.work_date else "",
            report.hours,
            "Техника" if grp == "техника" else "Ручная",
            tech_type,
            tech_name,
            activity,
            hand_work,
            report.location,
            report.crop,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    wb.save(filepath)
